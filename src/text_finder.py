# finder.py "example.xlsx" "aa bb cc"

from typing import Sequence
import openpyxl


class Cell:
    def __init__(self, coordinate, value):
        self.coordinate = coordinate
        self.value = value

class Finder:
    def __init__(self, path: str, sheet_name: str, targets: Sequence, find_from_header: bool):
        self.path = path
        self.sheet_name = sheet_name
        self.targets = targets
        self.find_from_header = find_from_header
        
        self.reset()


    def reset(self):
        self.is_finished = False;
        self.columns = []
        self.found_headers = []
        self.found_values = []


    def is_valid(self):
        if not self.path:
            return False
        
        if not self.sheet_name:
            return False

        if not self.targets:
            return False

        return True


    def run(self):
        if not self.is_valid():
            print('invalid params')
            return      
        
        self.reset()

        wb=None
        try:
            wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
            ws = wb[self.sheet_name]

            #dimension = ws.calculate_dimension()
            #print(dimension)
            
            for row in ws.iter_rows():
                if len(row) <= 0:
                    continue
                    
                if any(not c.value for c in row):
                    continue

                found_row = row
                break
                    
            if not found_row:
                return

            header_row = found_row[0].row
            for cell in found_row:
                self.columns.append(cell.value)

            if self.find_from_header:
                for col in self.columns:
                    if self.is_target(col):
                        self.found_headers.append(col)
                return
            else:
                start_row = header_row+1

                for row in ws.iter_rows(start_row):
                    for cell in row:
                        if self.is_target(str(cell.value)):
                            self.found_values.append(Cell(cell.coordinate, cell.value))

            self.is_finished = True
            return    

        finally:
            if wb:
                wb.close()


    def is_target(self, str_value:str):
        for target in self.targets:
            if str_value.find(target) >= 0 :
                return True
        
        return False
